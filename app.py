import os
import re
import io
import base64
import sys
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import webview
import chromadb
from sentence_transformers import SentenceTransformer
import datefinder

# Resolve database directory relative to app.py location
DB_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "chroma_db"))
COLLECTION_NAME = "quotation_items"

def clean_rate(val):
    """Parses raw pricing strings into numerical floats, filtering out dimension blocks and units."""
    if val is None:
        return 0.0
    val_str = str(val).strip()
    val_str_lower = val_str.lower()
    
    # Strip common currency names/symbols to avoid treating them as letters
    for cur in ['aed', 'dhs', 'dh', 'usd', 'aed.', 'dhs.']:
        val_str_lower = val_str_lower.replace(cur, '')
        
    val_str_lower = val_str_lower.strip()
    
    # Check if there are any remaining alphabetic characters
    # If there are, it is NOT a rate (e.g., dimension "L3 x W3.6m", unit "Qty", "Nos")
    if re.search(r'[a-zA-Z]', val_str_lower):
        return 0.0
        
    # Clean non-digit/non-punctuation characters
    val_str = re.sub(r'[^\d.,]', '', val_str)
    if not val_str:
        return 0.0
        
    # If string has both comma and dot (e.g. "33,600.00"), remove commas first
    if ',' in val_str and '.' in val_str:
        val_str = val_str.replace(',', '')
    # If string only has comma and 3 digits after (e.g. "33,600"), remove comma
    elif ',' in val_str:
        parts = val_str.split(',')
        if len(parts[-1]) == 3:
            val_str = val_str.replace(',', '')
        else:
            val_str = val_str.replace(',', '.')
            
    try:
        return float(val_str)
    except ValueError:
        return 0.0

class QuotationApi:
    def __init__(self):
        self.model = None
        self.client = chromadb.PersistentClient(path=DB_PATH)
        self.collection = self.client.get_or_create_collection(name=COLLECTION_NAME)
        self.sync_path = "G:\\My Drive"

    def _get_model(self):
        """Lazy loads sentence-transformers model to save initial window boot time."""
        if self.model is None:
            print("Loading sentence-transformers model (all-MiniLM-L6-v2)...")
            self.model = SentenceTransformer('all-MiniLM-L6-v2')
        return self.model

    def get_db_status(self):
        """Returns database status and number of indexed items."""
        try:
            count = self.collection.count()
            return {"status": "ready", "count": count}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def get_analytics(self):
        """Aggregates metadata stats about the active vector database index."""
        try:
            count = self.collection.count()
            if count == 0:
                return {
                    "total_items": 0,
                    "avg_price": "0.00",
                    "min_price": "0.00",
                    "max_price": "0.00",
                    "year_min": 2026,
                    "year_max": 2026
                }
                
            all_meta = self.collection.get(include=["metadatas"])
            rates = []
            years = []
            
            if all_meta and all_meta.get("metadatas"):
                for m in all_meta["metadatas"]:
                    r = float(m.get("historical_rate", 0.0))
                    if r > 0:
                        rates.append(r)
                    dt = m.get("quote_date", "")
                    if dt:
                        try:
                            yr = int(dt.split("-")[0])
                            years.append(yr)
                        except ValueError:
                            pass
                            
            total_items = count
            avg_price = sum(rates) / len(rates) if rates else 0.0
            min_price = min(rates) if rates else 0.0
            max_price = max(rates) if rates else 0.0
            year_min = min(years) if years else 2024
            year_max = max(years) if years else 2026
            
            return {
                "total_items": total_items,
                "avg_price": f"{avg_price:,.2f}",
                "min_price": f"{min_price:,.2f}",
                "max_price": f"{max_price:,.2f}",
                "year_min": year_min,
                "year_max": year_max
            }
        except Exception as e:
            print(f"Error computing database analytics: {e}")
            return {
                "total_items": 0,
                "avg_price": "0.00",
                "min_price": "0.00",
                "max_price": "0.00",
                "year_min": 2024,
                "year_max": 2026
            }

    def extract_date(self, filename, file_mtime):
        """Extracts date from filename using Regex and datefinder. Fallback to mtime."""
        match_dmy = re.search(r'(\d{1,2})[-_](\d{1,2})[-_](\d{4})', filename)
        if match_dmy:
            return f"{match_dmy.group(3)}-{match_dmy.group(2).zfill(2)}-{match_dmy.group(1).zfill(2)}"
            
        match_ymd = re.search(r'(\d{4})[-_](\d{1,2})[-_](\d{1,2})', filename)
        if match_ymd:
            return f"{match_ymd.group(1)}-{match_ymd.group(2).zfill(2)}-{match_ymd.group(3).zfill(2)}"
            
        try:
            matches = list(datefinder.find_dates(filename))
            if matches:
                return matches[0].strftime("%Y-%m-%d")
        except Exception:
            pass
            
        try:
            dt = datetime.fromtimestamp(file_mtime)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return "2025-01-01"

    def get_image_base64_thumbnail(self, img):
        """Extracts image data from openpyxl, handles BytesIO stream buffers, resizes, and encodes to base64."""
        try:
            pil_img = None
            raw_data = None
            
            # Check for file-like BytesIO objects inside img.ref
            if hasattr(img, 'ref') and img.ref:
                if hasattr(img.ref, 'read'):
                    try:
                        img.ref.seek(0)
                        raw_data = img.ref.read()
                    except Exception:
                        pass
                else:
                    pil_img = img.ref
            
            if pil_img is None and raw_data is None and hasattr(img, '_data') and img._data:
                try:
                    raw_data = img._data()
                except Exception:
                    pass
                    
            if raw_data:
                pil_img = PILImage.open(io.BytesIO(raw_data))
                
            if pil_img:
                # Copy and resize to thumbnail to save database space
                pil_img_copy = pil_img.copy()
                pil_img_copy.thumbnail((250, 250))
                
                buffered = io.BytesIO()
                pil_img_copy.save(buffered, format="PNG")
                img_str = base64.b64encode(buffered.getvalue()).decode('utf-8')
                return f"data:image/png;base64,{img_str}"
        except Exception as e:
            print(f"Error extracting image Base64: {e}")
        return ""

    def parse_excel_file(self, file_path):
        """Parses Excel sheets dynamically, identifying header alignments and mapping drawing images."""
        items = []
        path_obj = Path(file_path)
        file_name = path_obj.name
        
        try:
            mtime = path_obj.stat().st_mtime
            file_date = self.extract_date(file_name, mtime)
        except Exception:
            file_date = "2025-01-01"

        try:
            # data_only=True is CRITICAL to get calculated values instead of raw formula strings
            wb = openpyxl.load_workbook(str(path_obj), data_only=True)
        except Exception as e:
            print(f"Failed to open Excel '{file_name}': {e}")
            return items

        active_sheet_title = wb.active.title
        for sheet in wb.worksheets:
            # 1. DYNAMIC SHEET SELECTION
            sheet_title_lower = sheet.title.lower()
            should_parse = (
                sheet.title == active_sheet_title or
                any(x in sheet_title_lower for x in ["option", "opt", "sheet", "quotation", "production"])
            )
            if not should_parse:
                continue

            # 2. DYNAMIC COLUMN ALIGNMENT
            desc_col = None
            rate_col = None
            unit_col = None
            img_col = None
            header_row_idx = None
            
            # Scan top 15 rows for the header row containing the description cell
            for r_idx in range(1, 16):
                try:
                    row_vals = [sheet.cell(row=r_idx, column=c_idx).value for c_idx in range(1, 16)]
                except Exception:
                    continue
                row_str = [str(val).strip().lower() if val is not None else "" for val in row_vals]
                
                if any("description" in v or "particulars" in v for v in row_str):
                    header_row_idx = r_idx
                    for col_idx, val in enumerate(row_str, start=1):
                        if "description" in val or "particulars" in val:
                            desc_col = col_idx
                        elif "rate" in val or "price" in val:
                            rate_col = col_idx
                        elif "unit" in val:
                            unit_col = col_idx
                        elif "image" in val or "photo" in val:
                            img_col = col_idx
                    break

            # Fallback to standard columns if headers search fails
            if desc_col is None:
                desc_col = 2 # Column B
                img_col = 3  # Column C
                unit_col = 5 # Column E
                rate_col = 6 # Column F
                header_row_idx = 9

            # Map embedded product images to their rows (0-indexed anchor rows)
            images_by_row = {}
            if img_col and hasattr(sheet, '_images') and sheet._images:
                for img in sheet._images:
                    try:
                        col_idx = img.anchor._from.col
                        row_idx = img.anchor._from.row
                        # 3. MATCH IMAGES BY COLUMN INDEX
                        if col_idx == (img_col - 1):
                            images_by_row[row_idx + 1] = img
                    except Exception as e:
                        print(f"Error mapping image: {e}")

            # Parse data rows starting below header row
            start_row = (header_row_idx + 1) if header_row_idx else 10
            for r in range(start_row, sheet.max_row + 1):
                try:
                    desc_val = sheet.cell(row=r, column=desc_col).value if desc_col else None
                    rate_val = sheet.cell(row=r, column=rate_col).value if rate_col else None
                    unit_val = sheet.cell(row=r, column=unit_col).value if unit_col else None
                except Exception:
                    continue

                # STRICT VALIDATION: Check description is not blank
                if desc_val is None or str(desc_val).strip() == "":
                    continue
                    
                desc_str = str(desc_val).strip()
                
                # Skip total and summary rows
                if "total" in desc_str.lower() or "subtotal" in desc_str.lower():
                    continue

                # STRICT VALIDATION: Check Rate (AED) is not blank and is valid float
                rate_float = clean_rate(rate_val)
                if rate_float <= 0:
                    continue

                # Map Unit
                unit_str = "Qty"
                if unit_val is not None:
                    unit_str = str(unit_val).strip()

                # EXTRACT IMAGES VIA BASE64
                img_base64 = ""
                if r in images_by_row:
                    img_base64 = self.get_image_base64_thumbnail(images_by_row[r])

                items.append({
                    'original_description': desc_str,
                    'historical_rate': rate_float,
                    'unit': unit_str,
                    'quote_date': file_date,
                    'file_name': file_name,
                    'image_base64': img_base64
                })
                
        return items

    def parse_pdf_file(self, file_path):
        """4. MULTI-FORMAT (PDF Support) - Extract items using PyMuPDF (fitz) text parsing, scanning wide line ranges."""
        items = []
        path_obj = Path(file_path)
        file_name = path_obj.name
        
        try:
            mtime = path_obj.stat().st_mtime
            file_date = self.extract_date(file_name, mtime)
        except Exception:
            file_date = "2025-01-01"

        try:
            import fitz
            doc = fitz.open(str(path_obj))
        except Exception as e:
            print(f"Failed to open PDF '{file_name}': {e}")
            return items

        for page in doc:
            try:
                text = page.get_text("text")
                lines = [line.strip() for line in text.split('\n') if line.strip()]
            except Exception as e:
                print(f"Failed to extract text from PDF: {e}")
                continue
                
            idx = 0
            while idx < len(lines):
                line = lines[idx]
                
                # Filter out headers and footers
                if len(line) < 4 or any(x in line.lower() for x in ["page ", "quotation #", "date:", "client:", "donut", "boom tree"]):
                    idx += 1
                    continue
                    
                # Skip total and summary rows
                if "total" in line.lower() or "subtotal" in line.lower():
                    idx += 1
                    continue

                # Heuristic: Description text (non-numeric) followed by a price/rate
                if re.match(r'^\d+(\.\d+)?$', line):
                    idx += 1
                    continue
                    
                rate_found = None
                unit_found = "Qty"
                offset_used = 0
                
                # Scan up to next 15 lines to bypass detailed dimension lines and locate correct rates
                numbers = []
                for offset in range(1, 15):
                    if idx + offset < len(lines):
                        next_line = lines[idx + offset]
                        val = clean_rate(next_line)
                        if val > 0:
                            numbers.append((val, offset))
                
                # Determine correct rate vs. quantity (select first large number > 10, or fall back to last)
                if numbers:
                    large_numbers = [n for n in numbers if n[0] > 10.0]
                    if large_numbers:
                        rate_found = large_numbers[0][0]
                        offset_used = large_numbers[0][1]
                        
                        # Search for unit string in between description and rate
                        if offset_used > 1:
                            for u_offset in range(1, offset_used):
                                potential_unit = lines[idx + u_offset]
                                if len(potential_unit) < 8 and re.search(r'[a-zA-Z]', potential_unit):
                                    unit_found = potential_unit
                                    break
                    else:
                        rate_found = numbers[-1][0]
                        offset_used = numbers[-1][1]
                        
                if rate_found is not None:
                    items.append({
                        'original_description': line,
                        'historical_rate': rate_found,
                        'unit': unit_found,
                        'quote_date': file_date,
                        'file_name': file_name,
                        'image_base64': ""
                    })
                    idx += offset_used + 1
                else:
                    idx += 1
                    
        return items

    def index_files(self, path):
        """Indexes folder directories recursively using pathlib rglob for xlsx and pdf files."""
        path_obj = Path(path)
        if not path_obj.exists() or not path_obj.is_dir():
            return {"success": False, "error": f"Path '{path}' does not exist or is not a directory."}
            
        self.sync_path = str(path_obj)
        print(f"Recursively scanning target path: {self.sync_path}")
        
        try:
            excel_files = []
            pdf_files = []
            
            # 5. RECURSIVE DIRECTORY SCANNING
            for file in path_obj.rglob("*"):
                if file.name.startswith("~$") or not file.is_file():
                    continue
                file_lower = file.name.lower()
                if file.suffix.lower() == ".xlsx" and any(x in file_lower for x in ["quotation", "quote", "revised", "option", "production", "cost sheet", "cost estimate"]):
                    excel_files.append(file)
                elif file.suffix.lower() == ".pdf" and any(x in file_lower for x in ["quotation", "quote", "revised", "option", "production", "cost sheet", "cost estimate"]):
                    pdf_files.append(file)
                    
            all_items = []
            for file_path in excel_files:
                parsed_items = self.parse_excel_file(file_path)
                all_items.extend(parsed_items)
                
            for file_path in pdf_files:
                parsed_items = self.parse_pdf_file(file_path)
                all_items.extend(parsed_items)
                
            # Clear old database index
            try:
                self.client.delete_collection(name=COLLECTION_NAME)
            except Exception:
                pass
            self.collection = self.client.get_or_create_collection(name=COLLECTION_NAME)
            
            # Deduplicate items before embedding to avoid redundancy
            seen = set()
            unique_items = []
            for item in all_items:
                # Key on description and rate
                key = (item['original_description'].lower().strip(), item['historical_rate'])
                if key not in seen:
                    seen.add(key)
                    unique_items.append(item)
                    
            if not unique_items:
                return {
                    "success": True, 
                    "indexed_count": 0, 
                    "message": "Indexing Verified: 0 unique historical items from Donut (Blank Rows Filtered)"
                }
                
            descriptions = [item['original_description'] for item in unique_items]
            model = self._get_model()
            computed_embeddings = model.encode(descriptions, show_progress_bar=False)
            
            ids = []
            documents = []
            embeddings = []
            metadatas = []
            
            for idx, item in enumerate(unique_items):
                doc_id = f"item_{idx}"
                ids.append(doc_id)
                documents.append(item['original_description'])
                embeddings.append(computed_embeddings[idx].tolist())
                metadatas.append({
                    'original_description': item['original_description'],
                    'historical_rate': float(item['historical_rate']),
                    'unit': str(item['unit']),
                    'quote_date': str(item['quote_date']),
                    'file_name': str(item['file_name']),
                    'image_base64': str(item['image_base64'])
                })
                
            self.collection.add(
                ids=ids,
                embeddings=embeddings,
                metadatas=metadatas,
                documents=documents
            )
            
            return {
                "success": True,
                "indexed_count": len(unique_items),
                "message": f"Indexing Verified: {len(unique_items)} unique historical items from Donut (Blank Rows Filtered)"
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}

    def search_items(self, query, markup_rate=0.04):
        """Searches vector space and applies dynamic compound markup logic relative to 2026."""
        try:
            count = self.collection.count()
            if count == 0:
                return {"success": True, "matches": []}
                
            model = self._get_model()
            query_embedding = model.encode(query).tolist()
            
            results = self.collection.query(
                query_embeddings=[query_embedding],
                n_results=3
            )
            
            matches = []
            if results and results['ids'] and results['ids'][0]:
                ids = results['ids'][0]
                distances = results['distances'][0]
                metadatas = results['metadatas'][0]
                current_year = 2026
                
                # Clean markup rate
                markup_rate_val = float(markup_rate)
                
                for idx in range(len(ids)):
                    metadata = metadatas[idx]
                    distance = distances[idx]
                    similarity = 1.0 / (1.0 + distance)
                    
                    rate = float(metadata.get('historical_rate', 0.0))
                    quote_date = metadata.get('quote_date', '')
                    
                    historical_year = current_year
                    if quote_date:
                        try:
                            historical_year = int(quote_date.split('-')[0])
                        except ValueError:
                            pass
                            
                    elapsed_years = max(0, current_year - historical_year)
                    adjusted_rate = rate * ((1.0 + markup_rate_val) ** elapsed_years)
                    
                    matches.append({
                        'id': ids[idx],
                        'description': metadata.get('original_description', ''),
                        'original_rate': round(rate, 2),
                        'adjusted_rate': round(adjusted_rate, 2),
                        'unit': metadata.get('unit', 'Qty'),
                        'quote_date': quote_date,
                        'elapsed_years': elapsed_years,
                        'similarity': round(similarity * 100, 1),
                        'file_name': metadata.get('file_name', ''),
                        'image_base64': metadata.get('image_base64', '')
                    })
                    
            return {"success": True, "matches": matches}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def create_fallback_template(self, dest_path):
        """Creates standard blank Red Cube quotation format sheet from scratch."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Quotation"
        ws.views.sheetView[0].showGridLines = True
        
        navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        font_title = Font(name="Segoe UI", size=16, bold=True, color="1F497D")
        font_subtitle = Font(name="Segoe UI", size=11, italic=True)
        font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        font_bold = Font(name="Segoe UI", size=10, bold=True)
        font_normal = Font(name="Segoe UI", size=10)
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        double_bottom_border = Border(
            top=Side(style='thin', color='1F497D'),
            bottom=Side(style='double', color='1F497D')
        )
        
        ws['A2'] = "RED CUBE EVENT PRODUCTION"
        ws['A2'].font = font_title
        ws['A3'] = "QUOTATION SHEET"
        ws['A3'].font = font_subtitle
        
        ws['A5'] = "Client Name:"
        ws['A5'].font = font_bold
        ws['B5'] = ""
        ws['B5'].font = font_normal
        ws['E5'] = "Date:"
        ws['E5'].font = font_bold
        ws['F5'] = "=TODAY()"
        ws['F5'].font = font_normal
        
        headers = ["Item #", "Description", "Images", "Unit", "Qty", "Rate (AED)", "VAT 5%", "TOTAL (AED)"]
        for col_idx, text in enumerate(headers, start=1):
            cell = ws.cell(row=9, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = navy_fill
            cell.alignment = align_center
            cell.border = thin_border
            
        ws.row_dimensions[9].height = 25
        
        for r in range(10, 35):
            ws.row_dimensions[r].height = 55
            ws.cell(row=r, column=1, value=r-9).alignment = align_center
            ws.cell(row=r, column=1).font = font_normal
            ws.cell(row=r, column=1).border = thin_border
            
            ws.cell(row=r, column=2).border = thin_border
            ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            ws.cell(row=r, column=3).border = thin_border
            ws.cell(row=r, column=3).alignment = align_center
            
            ws.cell(row=r, column=4).border = thin_border
            ws.cell(row=r, column=4).alignment = align_center
            
            ws.cell(row=r, column=5).border = thin_border
            ws.cell(row=r, column=5).alignment = align_center
            
            ws.cell(row=r, column=6).border = thin_border
            ws.cell(row=r, column=6).alignment = align_right
            ws.cell(row=r, column=6).number_format = '#,##0.00'
            
            vat_cell = ws.cell(row=r, column=7, value=f"=ROUND(E{r}*F{r}*0.05, 2)")
            vat_cell.border = thin_border
            vat_cell.alignment = align_right
            vat_cell.number_format = '#,##0.00'
            
            tot_cell = ws.cell(row=r, column=8, value=f"=ROUND(E{r}*F{r}*1.05, 2)")
            tot_cell.border = thin_border
            tot_cell.alignment = align_right
            tot_cell.number_format = '#,##0.00'
            
        ws.cell(row=35, column=2, value="Grand Total (AED)").font = font_bold
        ws.cell(row=35, column=2).alignment = align_right
        
        sum_cell = ws.cell(row=35, column=8, value="=SUM(H10:H34)")
        sum_cell.font = font_bold
        sum_cell.border = double_bottom_border
        sum_cell.alignment = align_right
        sum_cell.number_format = '#,##0.00'
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 18
        
        wb.save(dest_path)

    def generate_excel(self, items, client_name):
        """Populates quote draft rows and outputs a compiled Excel quotation sheet."""
        try:
            sync_path_obj = Path(self.sync_path)
            
            template_path = sync_path_obj / "Red Cube - Quotation Format.xlsx"
            if not template_path.exists():
                template_path = Path("template.xlsx")
                if not template_path.exists():
                    self.create_fallback_template(str(template_path))
                    
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            output_filename = f"{client_name}_Quotation_{timestamp}.xlsx"
            
            try:
                output_path = sync_path_obj / output_filename
                with open(str(output_path), "w") as f:
                    pass
                os.remove(str(output_path))
            except Exception:
                output_path = Path(output_filename)
                
            wb = openpyxl.load_workbook(str(template_path), data_only=False)
            ws = wb.active
            
            header_row = 9
            col_map = {}
            for r in range(1, 25):
                row_vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
                row_str = [str(v).strip().lower() if v is not None else "" for v in row_vals]
                if any("description" in v or "particulars" in v for v in row_str):
                    header_row = r
                    for idx, val in enumerate(row_str, start=1):
                        if "item #" in val or "item no" in val or "s.no" in val or "sr no" in val:
                            col_map["item_num"] = idx
                        elif "description" in val or "particulars" in val:
                            col_map["description"] = idx
                        elif "images" in val or "image" in val or "photo" in val:
                            col_map["images"] = idx
                        elif "unit" in val:
                            col_map["unit"] = idx
                        elif "qty" in val or "quantity" in val:
                            col_map["qty"] = idx
                        elif "rate" in val or "rate (aed)" in val or "unit price" in val:
                            col_map["rate"] = idx
                        elif "vat" in val:
                            col_map["vat"] = idx
                        elif "total" in val or "total (aed)" in val:
                            col_map["total"] = idx
                    break
                    
            if "description" not in col_map:
                col_map = {
                    "item_num": 1, "description": 2, "images": 3, "unit": 4, 
                    "qty": 5, "rate": 6, "vat": 7, "total": 8
                }
                
            desc_col = col_map["description"]
            images_col = col_map.get("images", 3)
            unit_col = col_map.get("unit", 4)
            qty_col = col_map.get("qty", 5)
            rate_col = col_map.get("rate", 6)
            vat_col = col_map.get("vat", 7)
            tot_col = col_map.get("total", 8)
            
            if ws.cell(row=5, column=1).value and "client" in str(ws.cell(row=5, column=1).value).lower():
                ws.cell(row=5, column=2, value=client_name)
            ws.cell(row=5, column=6, value=datetime.now().strftime("%Y-%m-%d"))
            
            current_row = header_row + 1
            for idx, item in enumerate(items, 1):
                if "item_num" in col_map:
                    ws.cell(row=current_row, column=col_map["item_num"], value=idx)
                    
                ws.cell(row=current_row, column=desc_col, value=item['description'])
                ws.cell(row=current_row, column=unit_col, value=item['unit'])
                ws.cell(row=current_row, column=qty_col, value=float(item['qty']))
                ws.cell(row=current_row, column=rate_col, value=float(item['rate']))
                
                # Decode draft base64 image and place back inside generated quotation worksheet
                if item.get('image_base64') and images_col:
                    try:
                        b64_data = item['image_base64'].split(",")[1]
                        img_bytes = base64.b64decode(b64_data)
                        pil_img = PILImage.open(io.BytesIO(img_bytes))
                        
                        tmp_stream = io.BytesIO()
                        pil_img.save(tmp_stream, format="PNG")
                        tmp_stream.seek(0)
                        
                        oxl_img = openpyxl.drawing.image.Image(tmp_stream)
                        oxl_img.width = 70
                        oxl_img.height = 70
                        
                        cell_loc = f"{get_column_letter(images_col)}{current_row}"
                        ws.add_image(oxl_img, cell_loc)
                    except Exception as e:
                        print(f"Failed to insert image back into Excel: {e}")
                
                qty_letter = get_column_letter(qty_col)
                rate_letter = get_column_letter(rate_col)
                
                vat_cell = ws.cell(row=current_row, column=vat_col)
                if not vat_cell.value or not str(vat_cell.value).startswith('='):
                    vat_cell.value = f"=ROUND({qty_letter}{current_row}*{rate_letter}{current_row}*0.05, 2)"
                    
                tot_cell = ws.cell(row=current_row, column=tot_col)
                if not tot_cell.value or not str(tot_cell.value).startswith('='):
                    tot_cell.value = f"=ROUND({qty_letter}{current_row}*{rate_letter}{current_row}*1.05, 2)"
                    
                current_row += 1
                
            wb.save(str(output_path))
            return {
                "success": True, 
                "file_path": str(output_path.resolve()), 
                "filename": output_path.name
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}

def main():
    api = QuotationApi()
    window = webview.create_window(
        'Red Cube Smart Quotation Engine',
        'index.html',
        js_api=api,
        width=1180,
        height=820,
        resizable=True,
        background_color='#0a0a0a'
    )
    webview.start(debug=True)

if __name__ == '__main__':
    main()
