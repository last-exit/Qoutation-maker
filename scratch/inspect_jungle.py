import openpyxl
from pathlib import Path
import re

folder = Path(r"G:\My Drive\BOOM TREE\ALL THE QUOTATIONS")
print(f"Scanning {folder} for 'Jungle Gym'...")

for file in folder.glob("*.xlsx"):
    if file.name.startswith("~$"):
        continue
    try:
        wb = openpyxl.load_workbook(str(file), data_only=True)
    except Exception as e:
        continue
        
    for sheet in wb.worksheets:
        # Scan sheet for 'Jungle'
        found_rows = []
        for r in range(1, sheet.max_row + 1):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 20)]
            row_str = [str(v) if v is not None else "" for v in row_vals]
            for val in row_str:
                if "jungle gym" in val.lower():
                    found_rows.append((r, row_vals))
                    break
                    
        if found_rows:
            print("\n" + "="*50)
            print(f"FILE: {file.name}")
            print(f"SHEET: {sheet.title}")
            
            # Print top 15 rows to see headers
            print("--- HEADERS (FIRST 15 ROWS) ---")
            for r in range(1, 16):
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
                row_str = [str(v) if v is not None else "" for v in row_vals]
                if any("description" in v.lower() or "particulars" in v.lower() for v in row_str):
                    print(f"Row {r} (HEADER CANDIDATE): {row_vals}")
                else:
                    # Only print non-empty header candidate rows to be concise
                    if any(v for v in row_str):
                        print(f"Row {r}: {row_vals}")
                        
            print("--- MATCHED ROWS ---")
            for r, vals in found_rows:
                print(f"Row {r}: {vals}")
