"""Pulls every embedded image out of Excel, Word and PDF documents.

Three formats, three completely different burial sites for the same thing:

* **xlsx/xlsm** - openpyxl exposes drawings on `sheet._images`, anchored to a cell. The
  bytes live behind one of three different attributes depending on how the workbook was
  written, which is why `_store_openpyxl_image` tries all of them rather than assuming one.
* **docx** - a picture is an `<a:blip>` carrying a relationship id; the actual bytes are in
  a related part. Walking the body in document order is what gives each image a *location*
  (which table cell, which paragraph) rather than an anonymous pile. Headers, footers and
  floating shapes never appear in the body walk, so a second sweep over the package's image
  parts catches them and reports them as `unplaced`.
* **pdf** - PyMuPDF reports images per page with a bounding box. The box is kept: it is the
  only positional information a PDF gives, and it is what any downstream consumer would need
  to associate an image with the text beside it.

Nothing here interprets what an image *means*. It reports what the document contains, stores
the bytes once (see `store`), and records where each one came from.

One unreadable file never aborts a batch - it is reported in `skipped` and the rest still run.
"""
import io
from pathlib import Path

import store

SUPPORTED_EXCEL = {".xlsx", ".xlsm"}
SUPPORTED_WORD = {".docx"}
SUPPORTED_PDF = {".pdf"}
SUPPORTED = SUPPORTED_EXCEL | SUPPORTED_WORD | SUPPORTED_PDF

# Guard against a pathological page (a scanned document can hold hundreds of image
# fragments) locking the UI. The cap is reported in warnings, never silent.
MAX_IMAGES_PER_PDF_PAGE = 40


# --- Common helpers ---------------------------------------------------------------------

def _record(ref, source_path, kind, index, location, **extra):
    """Uniform record shape. `location` is the human label; the structured fields beside it
    are what a downstream consumer actually filters on."""
    record = {
        "ref": ref,
        "image_src": store.web_src(ref),
        "source_file": Path(source_path).name,
        "source_path": str(source_path),
        "kind": kind,
        "index": index,
        "location": location,
    }
    record.update(extra)
    return record


def _store_openpyxl_image(img):
    """Extracts image data from an openpyxl drawing and stores it. Returns a ref or "".

    openpyxl hands back the payload in one of three shapes depending on whether the workbook
    was loaded from disk or built in memory, so all three are tried before giving up.
    """
    try:
        raw_data = None

        if hasattr(img, "ref") and img.ref is not None:
            if hasattr(img.ref, "read"):
                try:
                    img.ref.seek(0)
                    raw_data = img.ref.read()
                except Exception:
                    pass
            elif hasattr(img.ref, "save"):
                buf = io.BytesIO()
                img.ref.save(buf, format="PNG")
                raw_data = buf.getvalue()

        if raw_data is None and hasattr(img, "_data") and img._data:
            try:
                raw_data = img._data()
            except Exception:
                pass

        if raw_data:
            return store.store_bytes(raw_data) or ""
    except Exception:
        return ""
    return ""


def store_raw_bytes(raw_bytes):
    """Stores bytes pulled straight out of a document part (docx/pdf). Returns a ref or ""."""
    return store.store_bytes(raw_bytes) or ""


# --- Excel ------------------------------------------------------------------------------

def _extract_excel(path, warnings):
    """Every embedded image on every worksheet, labelled with its anchor cell."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    # Not read-only: openpyxl discards drawings entirely in read-only mode, which is exactly
    # the data this tool exists to recover.
    workbook = openpyxl.load_workbook(str(path))
    images = []
    index = 0

    try:
        for sheet in workbook.worksheets:
            for img in getattr(sheet, "_images", None) or []:
                ref = _store_openpyxl_image(img)
                if not ref:
                    warnings.append(f"An image on sheet '{sheet.title}' could not be read.")
                    continue

                row = column = None
                cell = ""
                try:
                    # openpyxl anchors are 0-indexed; spreadsheet cells are 1-indexed.
                    row = img.anchor._from.row + 1
                    column = img.anchor._from.col + 1
                    cell = f"{get_column_letter(column)}{row}"
                except Exception:
                    # An image can be anchored absolutely (to the page rather than a cell).
                    # It is still a real image - report it without a cell rather than drop it.
                    pass

                index += 1
                images.append(_record(
                    ref, path, "xlsx", index,
                    f"{sheet.title}!{cell}" if cell else sheet.title,
                    sheet=sheet.title, cell=cell, row=row, column=column,
                ))
    finally:
        workbook.close()

    return images


# --- Word -------------------------------------------------------------------------------

def _cell_position(tc):
    """(table_number, row, column) for a `<w:tc>`, all 1-indexed, or (None, None, None)."""
    from docx.oxml.ns import qn

    try:
        tr = tc.getparent()
        tbl = tr.getparent()
        column = list(tr.findall(qn("w:tc"))).index(tc) + 1
        row = list(tbl.findall(qn("w:tr"))).index(tr) + 1
        body = tbl.getparent()
        table_no = list(body.findall(qn("w:tbl"))).index(tbl) + 1
        return table_no, row, column
    except Exception:
        return None, None, None


def _blip_location(body, blip):
    """Where in the document a picture sits: a table cell if it is in one, else a paragraph.

    Returns (placement, label, structured_fields).
    """
    from docx.oxml.ns import qn

    # A picture in a table cell is *also* inside a paragraph, so the whole ancestor chain is
    # walked before deciding: the cell is the more useful answer and must win.
    node = blip
    enclosing_paragraph = None
    while node is not None:
        tag = getattr(node, "tag", None)
        if tag == qn("w:tc"):
            table_no, row, column = _cell_position(node)
            label = (f"Table {table_no}, row {row}, col {column}"
                     if table_no else "Table cell")
            return "table-cell", label, {"table": table_no, "row": row, "column": column}
        if tag == qn("w:p") and enclosing_paragraph is None:
            enclosing_paragraph = node
        node = node.getparent()

    if enclosing_paragraph is not None:
        try:
            paragraph = list(body.iter(qn("w:p"))).index(enclosing_paragraph) + 1
        except Exception:
            paragraph = None
        return "body", (f"Paragraph {paragraph}" if paragraph else "Body"), {"paragraph": paragraph}

    return "body", "Body", {}


def _extract_word(path, warnings):
    """Body pictures in document order, then anything left in the package (headers,
    footers, floating shapes) reported as unplaced."""
    import docx
    from docx.oxml.ns import qn

    document = docx.Document(str(path))
    body = document.element.body
    images = []
    seen_parts = set()
    index = 0

    for blip in body.iter(qn("a:blip")):
        rid = blip.get(qn("r:embed")) or blip.get(qn("r:link"))
        if not rid:
            continue
        part = document.part.related_parts.get(rid)
        if part is None:
            continue

        ref = store_raw_bytes(getattr(part, "blob", None))
        if not ref:
            warnings.append(f"An image in '{Path(path).name}' could not be read.")
            continue
        seen_parts.add(id(part))

        placement, label, fields = _blip_location(body, blip)
        index += 1
        images.append(_record(ref, path, "docx", index, label,
                              placement=placement, **fields))

    # Headers, footers and floating shapes never show up in the body walk. They are still
    # images the document contains, so report them rather than pretending they do not exist.
    try:
        package_images = list(document.part.package.image_parts)
    except Exception:
        package_images = []

    for part in package_images:
        if id(part) in seen_parts:
            continue
        ref = store_raw_bytes(getattr(part, "blob", None))
        if not ref:
            continue
        index += 1
        images.append(_record(ref, path, "docx", index, "Unplaced (header/footer/shape)",
                              placement="unplaced"))

    return images


# --- PDF --------------------------------------------------------------------------------

def _extract_pdf(path, warnings):
    """Every embedded image on every page, with the bounding box it occupies.

    The box is the only positional information a PDF offers, and it is what lets a consumer
    line an image up with the text beside it - so it is recorded rather than discarded.
    """
    import fitz  # PyMuPDF

    doc = fitz.open(str(path))
    images = []
    index = 0

    try:
        for page_number, page in enumerate(doc, start=1):
            try:
                infos = page.get_image_info(xrefs=True)
            except Exception as exc:
                warnings.append(f"Page {page_number}: could not read image info ({exc}).")
                continue

            if len(infos) > MAX_IMAGES_PER_PDF_PAGE:
                warnings.append(
                    f"Page {page_number} holds {len(infos)} images; "
                    f"only the first {MAX_IMAGES_PER_PDF_PAGE} were taken."
                )
                infos = infos[:MAX_IMAGES_PER_PDF_PAGE]

            seen_xrefs = set()
            for info in infos:
                xref = info.get("xref")
                bbox = info.get("bbox")
                if not xref or not bbox or xref in seen_xrefs:
                    continue
                seen_xrefs.add(xref)
                try:
                    raw_bytes = doc.extract_image(xref).get("image")
                except Exception as exc:
                    warnings.append(f"Page {page_number}: image {xref} could not be read ({exc}).")
                    continue
                ref = store_raw_bytes(raw_bytes)
                if not ref:
                    warnings.append(f"Page {page_number}: image {xref} was not a readable image.")
                    continue

                index += 1
                images.append(_record(
                    ref, path, "pdf", index, f"Page {page_number}",
                    page=page_number, bbox=[round(float(v), 2) for v in bbox],
                    y0=round(float(bbox[1]), 2), y1=round(float(bbox[3]), 2),
                ))
    finally:
        doc.close()

    return images


# --- Public API ---------------------------------------------------------------------------

def extract_file(path):
    """Every embedded image in one document.

    Returns {"success", "file", "path", "kind", "images", "warnings"}. A document this tool
    cannot open comes back with success: False and a reason, never an exception.
    """
    path = Path(path)
    suffix = path.suffix.lower()

    if not path.exists():
        return {"success": False, "file": path.name, "path": str(path),
                "kind": "", "images": [], "warnings": [], "error": "File not found."}
    if suffix not in SUPPORTED:
        return {"success": False, "file": path.name, "path": str(path), "kind": "",
                "images": [], "warnings": [],
                "error": f"Unsupported type '{suffix}'. Accepts: {', '.join(sorted(SUPPORTED))}"}

    warnings = []
    kind = "xlsx" if suffix in SUPPORTED_EXCEL else ("docx" if suffix in SUPPORTED_WORD else "pdf")
    try:
        if suffix in SUPPORTED_EXCEL:
            images = _extract_excel(path, warnings)
        elif suffix in SUPPORTED_WORD:
            images = _extract_word(path, warnings)
        else:
            images = _extract_pdf(path, warnings)
    except Exception as exc:
        return {"success": False, "file": path.name, "path": str(path), "kind": kind,
                "images": [], "warnings": warnings, "error": str(exc)}

    return {"success": True, "file": path.name, "path": str(path), "kind": kind,
            "images": images, "warnings": warnings}


def extract_files(paths):
    """Runs `extract_file` over a batch.

    Returns {"success", "files", "images", "skipped", "warnings", "counts"}. `images` is the
    flat list across every document; `counts.unique` is how many distinct images that is once
    the content-addressed store has collapsed duplicates.
    """
    files, images, skipped, warnings = [], [], [], []

    for raw_path in paths or []:
        result = extract_file(raw_path)
        if not result["success"]:
            skipped.append({"file": result["file"], "path": result["path"],
                            "reason": result.get("error", "Could not be read.")})
            continue

        files.append({"file": result["file"], "path": result["path"], "kind": result["kind"],
                      "images": len(result["images"]), "warnings": result["warnings"]})
        images.extend(result["images"])
        warnings.extend(f"{result['file']}: {w}" for w in result["warnings"])

    unique_refs = {img["ref"] for img in images}
    return {
        "success": True,
        "files": files,
        "images": images,
        "skipped": skipped,
        "warnings": warnings,
        "counts": {
            "documents": len(files),
            "images": len(images),
            "unique": len(unique_refs),
            "skipped": len(skipped),
        },
    }
