"""Full pipeline against a built quotation archive: parse -> index -> search -> compile.

Runs entirely on temporary stores, so it never touches the live index, history or catalog.
This is the test that would have caught the destructive re-index and the positional ids,
because both only misbehave across a *second* sync.

The archive is built by `historical_archive` rather than read from `sample_quotes/`. What
remains in that folder is quotations this app *produced*, which `index_files` deliberately
skips as output rather than source pricing, so it indexes to nothing at all; the historical
client quotations that used to sit beside them were client data and are gone from the repo.
Reading the archive from disk was never the point of these tests — surviving a second sync is.
"""
from pathlib import Path

import pytest

HEADER_ROW = ["Description", "Unit", "Qty", "Unit Rate", "Total"]


def _add_priced_sheet(sheet, preamble, items):
    """Writes one worksheet in the layout `parsing.parse_excel_file` reads: a short header
    block, a row naming the columns, the priced lines, then a totals block.

    Every line's total agrees with its rate and quantity, so the parser can reconcile it and
    the item enters the index at high confidence rather than landing in the review queue.
    """
    for row in preamble:
        sheet.append(row)
    sheet.append(HEADER_ROW)
    for description, unit, qty, rate in items:
        sheet.append([description, unit, qty, rate, qty * rate])
    subtotal = sum(qty * rate for _, _, qty, rate in items)
    sheet.append(["Subtotal", None, None, None, subtotal])
    sheet.append(["Grand Total", None, None, None, round(subtotal * 1.05, 2)])


@pytest.fixture
def historical_archive(tmp_path):
    """A folder of historical quotations, in the shape the real archive had.

    Built rather than committed, both because the suite keeps no binary fixtures and because
    the quotations this stands in for were client data. Two files, one carrying a second
    priced option sheet, so multi-file and multi-sheet aggregation is exercised and ids have
    to stay stable across more than a single row. Venue is a labelled cell in one file and
    part of the filename in the other, covering both paths the extractor trusts.
    """
    import openpyxl

    folder = tmp_path / "historical"
    folder.mkdir()

    first = openpyxl.Workbook()
    first.active.title = "Quotation"
    _add_priced_sheet(
        first.active,
        preamble=[["Beachfront Activation - Quotation"], ["Ref", "HIST-2025-204"],
                  ["Venue", "Kite Beach"], []],
        items=[
            ("Main Stage Platform 8m x 6m", "Sqm", 1, 12500),
            ("Truss Roof Structure with Rigging", "Set", 1, 18000),
            ("LED Screen 4m x 3m", "Pcs", 2, 6500),
            ("Inflatable Bouncy Castle", "Pcs", 3, 850),
            ("Soft Play Area with Ball Pit", "Set", 1, 4200),
        ],
    )
    first.save(folder / "Beachfront Activation - Quotation - 09-12-2025.xlsx")

    second = openpyxl.Workbook()
    second.active.title = "Quotation"
    _add_priced_sheet(
        second.active,
        preamble=[["School Fair - Event Production"], ["Ref", "HIST-2024-118"], []],
        items=[
            ("Stage Riser 2m x 1m Section", "Pcs", 8, 450),
            ("Sound System with Mixing Desk", "Set", 1, 5500),
            ("Registration Desk with Branding", "Pcs", 2, 1200),
        ],
    )
    _add_priced_sheet(
        second.create_sheet("Option 2"),
        preamble=[["School Fair - Reduced Scope"], []],
        items=[
            ("Stage Riser 2m x 1m Section", "Pcs", 12, 420),
            ("Shade Structure 6m x 6m", "Set", 1, 7800),
        ],
    )
    second.save(folder / "School Fair @ Knowledge Village - 15-11-2024.xlsx")

    return folder


@pytest.fixture
def api(tmp_path, monkeypatch):
    """A QuotationApi wired to throwaway stores."""
    import app
    import catalog_db
    import corrections_db
    import db
    import history_db
    import image_store

    monkeypatch.setattr(app, "DB_PATH", str(tmp_path / "chroma"))
    monkeypatch.setattr(image_store, "IMAGE_DIR", tmp_path / "images")
    monkeypatch.setattr(db, "BACKUP_DIR", tmp_path / "backups")
    for module, name in ((history_db, "history.db"), (catalog_db, "catalog.db"),
                         (corrections_db, "corrections.db")):
        monkeypatch.setattr(module, "DB_FILE", str(tmp_path / name))
        module.init_db()

    instance = app.QuotationApi()
    instance.sync_path = str(tmp_path / "out")
    Path(instance.sync_path).mkdir(parents=True, exist_ok=True)
    return instance


def test_indexes_the_sample_archive(api, historical_archive):
    result = api.index_files(str(historical_archive))
    assert result["success"], result.get("error")
    assert result["indexed_count"] > 0
    assert api.collection.count() == result["indexed_count"]


def test_reindex_is_idempotent_and_ids_are_stable(api, historical_archive):
    """Ids must survive a re-sync — the review queue hands them to the UI and corrections
    write back to them."""
    api.index_files(str(historical_archive))
    first_ids = set(api.collection.get(include=[])["ids"])

    api.index_files(str(historical_archive))
    assert set(api.collection.get(include=[])["ids"]) == first_ids


def test_failed_reindex_leaves_the_previous_index_intact(api, historical_archive, monkeypatch):
    """The old code deleted the live collection before loading the model, so any failure
    afterwards wiped the company's entire price library."""
    api.index_files(str(historical_archive))
    before = api.collection.count()
    assert before > 0

    def explode(*args, **kwargs):
        raise RuntimeError("simulated embedding failure")

    monkeypatch.setattr(api._get_model(), "encode", explode)
    result = api.index_files(str(historical_archive))

    assert result["success"] is False
    assert api.collection.count() == before, "the price library must survive a failed sync"


@pytest.fixture
def archive_with_photo(tmp_path):
    """A one-row quotation workbook carrying a real embedded product photo.

    Built rather than committed so the suite has no binary fixtures, and kept separate from
    `historical_archive`, whose sheets are text-only — which is exactly how an image
    regression could pass unnoticed.
    """
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image

    folder = tmp_path / "archive"
    folder.mkdir()

    photo = tmp_path / "photo.png"
    Image.new("RGB", (300, 220), (180, 60, 40)).save(photo)

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["Description", "Unit", "Qty", "Rate", "Total"])
    sheet.append(["Pirate Ship", "Pcs", 2, 500, 1000])
    drawing = XLImage(str(photo))
    drawing.anchor = "F2"
    sheet.add_image(drawing)

    target = folder / "Cost Sheet - Test.xlsx"
    book.save(target)
    return folder


def test_index_stores_refs_not_image_bytes(api, archive_with_photo):
    import image_store

    result = api.index_files(str(archive_with_photo))
    assert result["success"], result.get("error")
    metadatas = api.collection.get(include=["metadatas"])["metadatas"]

    refs = [m.get("image_ref") for m in metadatas if m.get("image_ref")]
    assert refs, "the embedded photo should have been extracted"
    for ref in refs:
        assert image_store.is_ref(ref)
        assert image_store.exists(ref)
    assert not any("image_base64" in m for m in metadatas)


def test_extracted_photo_is_stored_once_on_disk(api, archive_with_photo):
    """Re-syncing the same archive must not accumulate duplicate image files."""
    import image_store

    api.index_files(str(archive_with_photo))
    after_first = image_store.stats()["count"]
    api.index_files(str(archive_with_photo))

    assert image_store.stats()["count"] == after_first


def test_search_returns_urls_not_base64(api, historical_archive):
    api.index_files(str(historical_archive))
    result = api.search_items("stage")
    assert result["success"]

    for match in result["matches"]:
        assert not match.get("image_src", "").startswith("data:")
        if match.get("image_ref"):
            assert match["image_src"].startswith("images/")


def test_compile_produces_a_document_and_a_history_row(api, historical_archive):
    import history_db

    api.index_files(str(historical_archive))
    result = api.compile_quotation({
        "client_name": "Test Client",
        "client_phone": "0501234567",
        "venue": "Kite Beach",
        "formats": ["xlsx"],
        "items": [{"description": "Pirate Ship", "unit": "Pcs", "qty": 2, "rate": 500.0}],
    })

    assert result["success"], result.get("error")
    assert Path(result["xlsx_path"]).exists()
    assert result["totals"]["subtotal"] == 1000.0
    assert result["totals"]["grand_total"] == 1050.0

    record = history_db.get_quotation_by_id(result["history_id"])
    assert record["quote_number"] == result["quote_ref"]
    assert record["client_id"] is not None


def test_compiled_document_embeds_the_photo(api, archive_with_photo):
    """A ref has to resolve all the way to bytes inside the generated file."""
    import zipfile

    api.index_files(str(archive_with_photo))
    ref = next(
        (m["image_ref"] for m in api.collection.get(include=["metadatas"])["metadatas"]
         if m.get("image_ref")),
        None,
    )
    assert ref, "expected at least one indexed photo"

    result = api.compile_quotation({
        "client_name": "Photo Client",
        "formats": ["xlsx"],
        "items": [{"description": "Pirate Ship", "unit": "Pcs", "qty": 1,
                   "rate": 100.0, "image_ref": ref}],
    })
    assert result["success"], result.get("error")
    assert result["image_failures"] == 0

    with zipfile.ZipFile(result["xlsx_path"]) as book:
        media = [n for n in book.namelist() if n.startswith("xl/media/")]
    assert media, "the photo should be embedded in the workbook"


def test_quote_numbers_do_not_repeat_across_compiles(api):
    refs = set()
    for i in range(3):
        result = api.compile_quotation({
            "client_name": f"Client {i}",
            "formats": ["xlsx"],
            "items": [{"description": "Item", "unit": "Pcs", "qty": 1, "rate": 10.0}],
        })
        assert result["success"], result.get("error")
        refs.add(result["quote_ref"])
    assert len(refs) == 3


def test_costs_attach_from_the_catalog_for_margin_reporting(api):
    """The exact-string lookup this replaces never matched a real multi-line quote line, so
    every quote recorded cost_price: None and margin reporting had no data."""
    import catalog_db

    catalog_db.add_catalog_item("Pirate Ship", rate=500, cost_price=300)
    enriched = api._attach_costs([
        {"description": "Pirate Ship\nL5.8 x W3.4 x H3.2m\n- Twin swings", "qty": 1, "rate": 500},
        {"description": "Helicopter Charter", "qty": 1, "rate": 900},
    ])

    assert enriched[0]["cost_price"] == 300
    assert enriched[1]["cost_price"] is None


def test_venue_correction_survives_resync_without_freezing_the_rate(api, archive_with_photo):
    """The two bugs that made the review queue untrustworthy, checked together: the
    correction must persist across a sync, and it must not pin the price."""
    import corrections_db

    api.index_files(str(archive_with_photo))
    queue = api.get_review_queue()
    assert queue["success"] and queue["items"], "an unlabelled venue should queue for review"

    target = queue["items"][0]
    assert api.save_correction(target["id"], venue="Kite Beach")["corrected_fields"] == ["venue"]

    stored = corrections_db.get_correction(target["file_name"], target["description"])
    assert stored["corrected_fields"] == "venue", "only the edited field may be pinned"

    api.index_files(str(archive_with_photo))
    after = api.collection.get(ids=[target["id"]], include=["metadatas"])["metadatas"][0]
    assert after["venue"] == "Kite Beach", "the correction should survive the re-sync"
    assert after["historical_rate"] == target["rate"], "the rate must still come from source"


def test_reprice_in_the_source_reaches_the_app_after_a_venue_correction(api, archive_with_photo):
    """The headline symptom of the old behaviour: once anyone touched an item's venue, a
    price change in the source spreadsheet could never reach the app again."""
    import openpyxl

    api.index_files(str(archive_with_photo))
    target = api.get_review_queue()["items"][0]
    api.save_correction(target["id"], venue="Kite Beach")

    source = next(Path(archive_with_photo).glob("*.xlsx"))
    book = openpyxl.load_workbook(source)
    sheet = book.active
    sheet["D2"] = 750          # was 500
    sheet["E2"] = 1500
    book.save(source)

    api.index_files(str(archive_with_photo))
    rates = [m["historical_rate"] for m in api.collection.get(include=["metadatas"])["metadatas"]]
    assert 750 in rates, "the new price must reach the index"


def test_dismissal_is_not_re_flagged_on_the_next_sync(api, archive_with_photo):
    api.index_files(str(archive_with_photo))
    target = api.get_review_queue()["items"][0]
    assert api.dismiss_review_item(target["id"])["success"]

    api.index_files(str(archive_with_photo))
    still_queued = {i["id"] for i in api.get_review_queue()["items"]}
    assert target["id"] not in still_queued


def test_maintenance_leaves_the_index_queryable(api, historical_archive):
    """`count()` is not proof the index survived.

    Truncating Chroma's embeddings_queue reclaimed 57 MB and left count() reporting every
    row, while every get() and query() failed with "Error finding id" — the HNSW segment
    replays from that log. Maintenance must therefore be checked by actually reading.
    """
    api.index_files(str(historical_archive))
    before = api.collection.count()

    result = api.run_maintenance()
    assert result["success"], result.get("error")
    assert api.collection.count() == before
    assert all(v == "ok" for v in result["summary"]["integrity"].values())

    fetched = api.collection.get(limit=3, include=["metadatas"])
    assert fetched["ids"], "ids must still resolve after maintenance"

    search = api.search_items("stage")
    assert search["success"] and search["matches"], "vector search must still return hits"


def test_orphan_sweep_refuses_an_empty_reference_set(api, archive_with_photo):
    """An empty live set means the caller failed to read, not that every photo is garbage.

    Indexed against the archive that carries a photo, so there is something on disk the
    sweep could wrongly delete — against a text-only archive it would pass on an empty store.
    """
    import maintenance

    api.index_files(str(archive_with_photo))
    assert maintenance.collect_image_orphans(set()) == []
